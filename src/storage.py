"""Storage layer for tiktok-hashtag-collector.

Handles CSV/Excel output, buffer management, deduplication integration,
and date-based file rotation for monitor mode.
"""

from __future__ import annotations

import threading
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.config import AppConfig
from src.dedup import DedupStore
from src.logger import get_logger
from src.models import VideoRecord
from src.utils import sanitize_csv_cell, utcnow_naive

logger = get_logger()

_URL_COLUMN: str = "url"


class StorageManager:
    """Manages buffered writes to CSV/Excel output files with deduplication.

    Thread-safe via an internal write lock. Records are accumulated in a
    buffer and flushed when the buffer reaches ``config.write_batch_size``
    or when ``flush()`` is called explicitly.
    """

    def __init__(self, config: AppConfig, dedup: DedupStore) -> None:
        self._config = config
        self._dedup = dedup
        self._write_lock: threading.Lock = threading.Lock()
        self._buffer: list[VideoRecord] = []
        self._current_date: str = ""  # tracks date for file rotation in monitor mode

    @property
    def dedup(self) -> DedupStore:
        """Public read-only accessor for the underlying dedup store."""
        return self._dedup

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def get_output_path(self, hashtag: str, combined: bool = False) -> Path:
        """Return the output file path for *hashtag*.

        Creates ``config.output_dir`` if it does not already exist.

        Args:
            hashtag: The source hashtag (used in the filename when not combined).
            combined: When ``True`` returns a ``combined_<date>`` filename.

        Returns:
            A :class:`~pathlib.Path` with ``.csv`` extension (or ``.xlsx`` when
            ``config.output_format == "excel"``).
        """
        self._config.output_dir.mkdir(parents=True, exist_ok=True)
        date: str = utcnow_naive().strftime("%Y-%m-%d")
        stem: str = f"combined_{date}" if combined else f"{hashtag}_{date}"
        suffix: str = ".xlsx" if self._config.output_format == "excel" else ".csv"
        return self._config.output_dir / f"{stem}{suffix}"

    def load_existing(self, hashtag: str, combined: bool = False) -> set[str]:
        """Read already-written video IDs from the output file on disk.

        Used to pre-populate the dedup store at startup so that re-runs
        of the scraper do not produce duplicate rows.

        Args:
            hashtag: Source hashtag — forwarded to :meth:`get_output_path`.
            combined: Forwarded to :meth:`get_output_path`.

        Returns:
            A set of video ID strings found in the existing file, or an
            empty set when the file does not exist or is unreadable.
        """
        # When format is csv/excel a single path is returned; for "both" the
        # CSV is the canonical source (it is always written and never read by
        # openpyxl in append mode), so prefer it whenever it exists.
        primary: Path = self.get_output_path(hashtag, combined)
        csv_path: Path = primary if primary.suffix == ".csv" else primary.with_suffix(".csv")
        xlsx_path: Path = primary if primary.suffix == ".xlsx" else primary.with_suffix(".xlsx")

        try:
            if csv_path.exists():
                df = pd.read_csv(csv_path, usecols=["video_id"], dtype={"video_id": str})
                return set(df["video_id"].dropna().unique())
            if xlsx_path.exists():
                df = pd.read_excel(xlsx_path, usecols=["video_id"], dtype={"video_id": str})
                return set(df["video_id"].dropna().unique())
        except (FileNotFoundError, KeyError, ValueError, pd.errors.EmptyDataError):
            pass
        return set()

    def write_records(
        self,
        records: list[VideoRecord],
        hashtag: str,
        combined: bool = False,
    ) -> int:
        """Add *records* to the write buffer, flushing when full.

        Only records that pass deduplication are buffered. Each accepted
        record is immediately marked as seen in the dedup store so that
        concurrent callers cannot double-write the same video.

        Args:
            records: Candidate records to persist.
            hashtag: Source hashtag used for output path resolution.
            combined: Forwarded to :meth:`get_output_path` / :meth:`flush`.

        Returns:
            The number of new (non-duplicate) records accepted into the buffer.
            Does **not** equal the number of records flushed to disk; call
            :meth:`flush` after your scraping loop to drain the remaining buffer.
        """
        with self._write_lock:
            new_records: list[VideoRecord] = self._dedup.filter_new(records)
            if not new_records:
                return 0

            self._buffer.extend(new_records)

            for record in new_records:
                self._dedup.mark_seen(record.video_id)

            if len(self._buffer) >= self._config.write_batch_size:
                self._flush_locked(hashtag, combined)

            return len(new_records)

    def flush(self, hashtag: str, combined: bool = False) -> int:
        """Flush all buffered records to disk immediately.

        Thread-safe wrapper around :meth:`_flush_locked`.

        Args:
            hashtag: Source hashtag used for output path resolution.
            combined: Forwarded to :meth:`get_output_path`.

        Returns:
            The number of records written in this flush.
        """
        with self._write_lock:
            return self._flush_locked(hashtag, combined)

    def stats(self) -> dict:
        """Return a snapshot of buffer and deduplication statistics.

        Returns:
            A dict with keys ``buffer_size`` and ``dedup_stats``.
        """
        return {
            "buffer_size": len(self._buffer),
            "dedup_stats": self._dedup.stats,
        }

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _flush_locked(self, hashtag: str, combined: bool = False) -> int:
        """Write buffered records to disk. Caller must hold ``_write_lock``.

        Handles date-based path rotation: if the UTC date has advanced since
        the last flush the path is recalculated so records land in the correct
        daily file. Appends to an existing file or creates a new one with a
        header row. Optionally writes an Excel copy when
        ``config.output_format`` is ``"excel"`` or ``"both"``.

        Args:
            hashtag: Source hashtag used for output path resolution.
            combined: Forwarded to :meth:`get_output_path`.

        Returns:
            The number of records written, or ``0`` when the buffer was empty.
        """
        if not self._buffer:
            return 0

        count: int = len(self._buffer)
        df: pd.DataFrame = pd.DataFrame([r.to_dict() for r in self._buffer])

        # Defuse spreadsheet formula injection across every string column.
        # Scraped fields like description / author_username / music_title can
        # contain payloads such as "=cmd|'/c calc'!A0" that execute on open.
        for column in df.select_dtypes(include="object").columns:
            df[column] = df[column].map(sanitize_csv_cell)

        # Date-rotation: recalculate path each flush (get_output_path uses today's UTC date)
        today: str = utcnow_naive().strftime("%Y-%m-%d")
        if self._current_date and self._current_date != today:
            logger.info(
                "Date rolled over from %s → %s; new records go to a fresh daily file",
                self._current_date,
                today,
            )
        self._current_date = today

        path: Path = self.get_output_path(hashtag, combined)

        # CSV write for "csv" / "both"
        if self._config.output_format in ("csv", "both"):
            csv_path: Path = path if path.suffix == ".csv" else path.with_suffix(".csv")
            self._write_csv(df, csv_path)

        # Excel write for "excel" / "both"
        if self._config.output_format in ("excel", "both"):
            xlsx_path: Path = path if path.suffix == ".xlsx" else path.with_suffix(".xlsx")
            self._write_excel(df, xlsx_path, append=xlsx_path.exists())

        self._buffer.clear()
        logger.debug(
            f"Flushed {count} records to {path}",
            extra={"hashtag": hashtag, "count": count},
        )
        return count

    def _write_csv(self, df: pd.DataFrame, path: Path) -> None:
        """Append *df* to *path* as CSV, writing a header only for new files.

        Args:
            df: DataFrame containing records to write.
            path: Target CSV file path.
        """
        file_exists: bool = path.exists()
        df.to_csv(
            path,
            mode="a" if file_exists else "w",
            header=not file_exists,
            index=False,
            encoding="utf-8-sig",
        )

    def _write_excel(
        self,
        df: pd.DataFrame,
        path: Path,
        append: bool,
    ) -> None:
        """Write *df* to an Excel file at *path* with formatting.

        When *append* is ``True`` and the file exists, new rows are appended
        in-place via openpyxl (O(new_rows)) instead of reading+concatenating
        the full sheet (O(total_rows)). A fresh file gets the full formatting
        pass; appends only format the freshly added rows (header already set).

        Args:
            df: DataFrame rows to persist.
            path: Destination ``.xlsx`` path.
            append: When ``True``, merge with the file's existing content.
        """
        if append and path.exists():
            try:
                wb = load_workbook(path)
                ws = wb.active
                url_col_index: int | None = None
                try:
                    url_col_index = list(df.columns).index(_URL_COLUMN) + 1
                except ValueError:
                    pass
                for row in df.itertuples(index=False, name=None):
                    ws.append(list(row))
                    if url_col_index is not None:
                        cell = ws.cell(row=ws.max_row, column=url_col_index)
                        url_value = str(cell.value or "")
                        if url_value.startswith("http"):
                            cell.hyperlink = url_value
                            cell.font = Font(color="0563C1", underline="single")
                wb.save(path)
                return
            except (FileNotFoundError, ValueError, KeyError):
                pass  # fall through to full rewrite

        df.to_excel(path, index=False, engine="openpyxl")
        self._format_excel(df, path)

    def _format_excel(self, df: pd.DataFrame, path: Path) -> None:
        """Apply openpyxl formatting to *path* in-place.

        Formatting applied:
        - Bold font on the header row (row 1)
        - Freeze pane at cell A2 so the header stays visible on scroll
        - Auto-column width clamped to [10, 60] characters
        - Hyperlinks on every cell in the ``url`` column

        Args:
            df: The DataFrame that was written (used for column metadata).
            path: Path to the ``.xlsx`` file to format.
        """
        wb = load_workbook(path)
        ws = wb.active

        # Bold header
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto column width and hyperlinks
        url_col_index: int | None = None
        try:
            url_col_index = df.columns.get_loc(_URL_COLUMN) + 1  # 1-based
        except KeyError:
            pass

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            col_letter: str = get_column_letter(col_idx)
            max_length: int = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
                    if url_col_index is not None and col_idx == url_col_index:
                        url_value: str = str(cell.value)
                        if url_value.startswith("http"):
                            cell.hyperlink = url_value
                            cell.font = Font(color="0563C1", underline="single")

            adjusted_width: int = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(path)
